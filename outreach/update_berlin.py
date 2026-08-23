import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
ws = wb['Berlin (Pilot)']

updates = {
    'BERLIN Souvenir-Store - The Wall Store': (
        'https://www.souvenir-shop-berlin.de/', 'B2B wholesale portal: souvenir-shop-berlin.de/B2B-Grosshandel/ (form) — dedicated souvenir wholesale/shelf-service program', 'Found',
        'Chain "Hauptstadtmarke BERLIN" (op. by same group as The Wall Store/BERLIN Souvenir-Store, 2000m² warehouse, 48h delivery). Reuse for DomAquarée location.'),
    'Souvenirshop': (
        'https://muddastadt-berlin.de/store/souvenirs/', 'info@muddastadt-berlin.de', 'Found',
        'Likely same MUDDASTADT GmbH parent as Geschenkartikel BERLIN STORE/I LOVE BERLIN entries above — reuse B2B trade shop shop.muddastadt-berlin.de'),
    'Berlin Wallstore Souvenirs': (
        'https://www.pawlowski-souvenirs.de/', 'Contact form: pawlowski-souvenirs.de/index.php/kontaktieren — no public email found (site blocked direct fetch)', 'Partial',
        'Independent souvenir shop since 1992, sells Wall/Ampelmann items.'),
    'BERLIN Souvenir-Store - DomAquarée': (
        'https://www.souvenir-shop-berlin.de/', 'B2B wholesale portal: souvenir-shop-berlin.de/B2B-Grosshandel/ (form)', 'Found',
        'Same chain/operator as Wall Store Mercedes/Uber Platz location above — reuse contact.'),
    'Berlins Souvenirs x Kiosk': (
        'https://www.berlins-souvenirs.de/', 'Contact form via berlins-souvenirs.de — no confirmed email (site combines souvenir + UPS parcel point)', 'Partial',
        'Small kiosk-format shop; verify in person it is not primarily a parcel/UPS point before pitching.'),
    'Denns BioMarkt (Hauptbahnhof)': (
        'https://www.biomarkt.de/', 'info@denns.de (new-product/category management contact per dennree)', 'Found',
        'National chain (dennree Biohandelshaus) — HQ decides listings, not individual stores. Reuse info@denns.de for all Denns locations in other cities.'),
    'Der Bioladen': (
        'TBD (no independent website found; Facebook page only)', 'marggraf.m@gmx.de (owner, per directory listings)', 'Partial',
        'Small independent shop with café; no dedicated business website found — email sourced from directory listing, not a verified Impressum.'),
    'Biogoods – Bioladen, Feinkost & Kaffeebar': (
        'https://www.biogoods.de/', 'hallo@biogoods.de', 'Found', None),
    'Kollektiv Bioase': (
        'https://bioase.berlin/', 'hallo@bioase.berlin (contact form also at bioase.berlin/contact/)', 'Found',
        'Worker-collective organic market/café — note cooperative structure in pitch.'),
    'VITALIA Reformhaus (Friedrichstraße)': (
        'https://www.vitalia-reformhaus.de/', 'shop@vitalia-reformhaus.de (HQ, Bruckmühl) — contact page: vitalia-reformhaus.de/kontakt', 'Partial',
        'Chain HQ (VITALIA GmbH, Bruckmühl, Bavaria) — no dedicated B2B/vertrieb email found via search; shop@ is the online-shop contact. Reuse for all VITALIA Berlin locations. Worth a follow-up call to confirm right buyer contact.'),
    'Reformhaus Quentin GmbH (Grunerstraße)': (
        'https://www.reformhaus-quentin.de/', 'gscholz@reformhaus-quentin.de', 'Found',
        'Berlin-based chain (HRB 136237 B). Reuse for both Quentin locations.'),
    'VITALIA Reformhaus (Kaiser-Wilhelm-Passage)': (
        'https://www.vitalia-reformhaus.de/', 'shop@vitalia-reformhaus.de (HQ)', 'Partial',
        'Same VITALIA chain as Friedrichstraße location — reuse contact.'),
    'Reformhaus Quentin GmbH (Tempelhofer Damm)': (
        'https://www.reformhaus-quentin.de/', 'gscholz@reformhaus-quentin.de', 'Found',
        'Same Quentin chain as Grunerstraße location — reuse contact.'),
    'Reformhaus DEMSKI': (
        'https://www.demski.de/', 'info@demski.de (HQ, Volkmarstraße) / rs@demski.de (this branch direct)', 'Found',
        'Berlin chain with multiple filialen — HQ info@demski.de is the general/wholesale-appropriate contact; reuse for other Demski branches if found in later cities.'),
    'Reformhaus (Wilmersdorfer Str.)': (
        'https://www.reformhaus.de/', 'TBD — no dedicated shop website found; independently run by owner Bagheri', 'Not started',
        'Listed as "Reformhaus Bagheri" in directories — appears independent, not part of a chain. Phone/fax only found; needs a follow-up search or phone call for email.'),
    'VITALIA Reformhaus (Rüdesheimer Str.)': (
        'https://www.vitalia-reformhaus.de/', 'shop@vitalia-reformhaus.de (HQ)', 'Partial',
        'Same VITALIA chain — reuse contact.'),
    'VITALIA Reformhaus (Schönhauser Allee)': (
        'https://www.vitalia-reformhaus.de/', 'shop@vitalia-reformhaus.de (HQ)', 'Partial',
        'Same VITALIA chain — reuse contact.'),
    'Reformhaus Höfeler': (
        'https://reformhaus-hoefeler.jimdofree.com/', 'reformhaus-hoefeler@gmx.net', 'Found',
        'Independent, family-run (2 Berlin locations).'),
    'MediosApotheke Oranienburger Tor': (
        'https://www.mediosapotheke.de/standorte/oranienburgertor/', 'oranienburger-tor@mediosapotheke.de', 'Found',
        'Chain (4 Berlin locations) — specialty pharmacy, 24h. Note: Apotheken in Germany face strict rules on unsolicited product promotion (Heilmittelwerbegesetz) in addition to UWG/GDPR — vet approach carefully.'),
    'Sanimedius pharmacy Rosenthaler Platz': (
        'https://www.sanimedius.de/rosenthaler/', 'rosenthalerplatz@sanimedius.de (alt: mail@apotheke-rosenthalerplatz.de)', 'Found', None),
    'pharmacy Hackesches Quartier': (
        'https://www.apotheke-hackesches-quartier.de/', 'info@apotheke-hackesches-quartier.de', 'Found', None),
    'Arkona Apotheke': (
        'https://www.arkona-apotheke.de/', 'info@arkona-apotheke.de', 'Found', None),
}

count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    name = row[2].value
    if name in updates:
        website, contact, status, note = updates[name]
        row[6].value = website
        row[7].value = contact
        row[8].value = status
        if note:
            row[9].value = note
        count += 1

print('updated rows:', count)
wb.save('ginger_shoc_outreach_tracker.xlsx')
