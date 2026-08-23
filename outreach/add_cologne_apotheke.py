import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
ws = wb['Cologne']

# (name, website)
rows = [
('Kosmas-Apotheke', 'TBD'),
('Stadt-Apotheke', 'TBD'),
('DocVerde Apotheke Köln-Mülheim OHG', 'TBD'),
('Alte Apotheke in Junkersdorf', 'TBD'),
('Max und Moritz Apotheke', 'TBD'),
('Jakobus-Apotheke', 'TBD'),
('Böcking-Apotheke', 'TBD'),
('Wildpark-Apotheke', 'TBD'),
('Flora-Apotheke', 'TBD'),
('Kastell-Apotheke', 'TBD'),
('Lindenthal-Apotheke', 'https://www.lindenthal-apotheke.de/'),
('Rathaus Apotheke Rodenkirchen', 'TBD'),
('Rhein-Apotheke (Rodenkirchen)', 'https://www.rhein-apo.de/'),
('Markt-Apotheke Porz', 'http://www.apothekeporz.de/'),
('Damian-Apotheke', 'http://www.apothekeporz.de/'),
('Marien-Apotheke (Porz)', 'http://www.apothekeporz.de/'),
('Bayenthal-Apotheke', 'TBD'),
('Goltstein Apotheke', 'TBD'),
('Zollstock-Apotheke', 'TBD'),
('Marienburg Apotheke', 'https://www.marienburg-apotheke-koeln.de/'),
('Buchheimer-Apotheke', 'TBD'),
('Burg Apotheke Köln', 'https://burg-apotheke-koeln.de/'),
('Alpha-Apotheke', 'https://www.alpha-apotheke-koeln.de/'),
('easyApotheke Sülzburgstraße', 'https://suelzburgstrasse.easyapotheken.de/'),
('Sülzburg-Apotheke', 'https://suelzburg-apo-koeln.de/'),
('Apotheke am Questerhof', 'TBD'),
]

for name, website in rows:
    ws.append(('Cologne','Apotheke',name,'TBD','TBD','TBD',website,'TBD','Not started',None))

wb.save('ginger_shoc_outreach_tracker.xlsx')
count=0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]=='Apotheke': count+=1
print('Cologne Apotheke total:', count)
