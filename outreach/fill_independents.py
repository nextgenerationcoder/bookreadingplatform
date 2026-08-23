import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')

# Fill contact info for existing independent rows still marked Not started
fills = {
    ('Leipzig','Reformhaus Sonntag'): dict(phone='+49 341 6886677', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Leipzig','Reformhaus Blicke (Ritterstraße)'): dict(phone='+49 341 9605964', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Leipzig','Reformhaus Blicke (Friedrich-Ebert-Str.)'): dict(phone='+49 341 2536203', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Dortmund','Reformhaus Kirchlinde (A. Ogniwek u. I. Rothardt)'): dict(phone='+49 231 7267228', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Essen','Steeler Reformhaus'): dict(phone='+49 201 513486', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Bremen','Reformhaus Bühring'): dict(phone='+49 421 662118', website='http://reformhaus-buehring.de/', contact='info@reformhaus-buehring.de', status='Found'),
    ('Dresden','Reformhaus Ferstl'): dict(phone='+49 351 3144817', website='TBD — no dedicated website found', contact='TBD — no public email found; phone contact only', status='Not started'),
    ('Dresden','Reformhaus Mewald'): dict(phone='+49 351 6494048', website='TBD — no dedicated website found', contact='TBD — no public email found (in Altmarkt-Galerie mall); phone contact only', status='Not started'),
    ('Dresden','Reformhaus Langner'): dict(phone='+49 351 2153427', website='http://www.reformhaus-langner.de/', contact='Contact form via reformhaus-langner.de — no public email found', status='Partial'),
}

for (sheet, name), u in fills.items():
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[2].value == name:
            row[4].value = u['phone']
            row[6].value = u['website']
            row[7].value = u['contact']
            row[8].value = u['status']
            break

# Add newly found independents
new_rows = {
    'Stuttgart': [
        ('Stuttgart','Reformhaus','Grünschnabel Naturkost','Sigmundtstr. 1, 70563 Stuttgart-Vaihingen','+49 711 7352502','TBD','TBD — no dedicated modern website found (small local site)','naturkost-gruenschnabel@arcor.de','Found','Independent, family-run since 1983.'),
    ],
    'Düsseldorf': [
        ('Düsseldorf','Reformhaus','Reformhaus Christiane Worthoff (Eller)','Auf\'m Großenfeld 5, 40229 Düsseldorf-Eller','+49 211 211346','TBD','TBD — no dedicated website found','TBD — no public email found; phone contact only','Not started','Independent, single owner-operator.'),
    ],
    'Hannover': [
        ('Hannover','Reformhaus','naturahaus (Ute Schneider e.K.)','Lister Meile 88, 30161 Hannover','+49 511 665942','TBD','http://naturahaus.net/','info@naturahaus-hannover.de','Found','Independent single-location shop (owner Ute Schneider); listed in Vita Nova cooperative directory for marketing/branding purposes but is its own company, not a multi-branch chain — verify by phone that purchasing is locally decided before pitching as fully independent.'),
    ],
}

for sheet, rows in new_rows.items():
    ws = wb[sheet]
    for r in rows:
        ws.append(r)

wb.save('ginger_shoc_outreach_tracker.xlsx')
print('done')
