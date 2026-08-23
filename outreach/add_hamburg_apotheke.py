import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
ws = wb['Hamburg']

rows = [
('Hamburg','Apotheke','Apotheke Innenstadt (Rödingsmarkt)','Rödingsmarkt 1, 20459 Hamburg','+49 40 378673','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Innenstadt (Spitalerstraße)','Spitalerstraße 7, 20095 Hamburg','+49 40 32526271','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Hamburg-Altstadt','Jacobikirchhof 8, 20095 Hamburg','+49 40 335190','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Innenstadt (Bergstraße)','Bergstraße 14, 20095 Hamburg','+49 40 32527690','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Innenstadt (Colonnaden)','Colonnaden 36, 20354 Hamburg','+49 40 346566','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Innenstadt (Steindamm)','Steindamm 32, 20099 Hamburg','+49 40 245350','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Die Pfeil Apotheke','Eimsbütteler Chaussee 28, 20259 Hamburg','+49 40 4390025','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','VITA Apotheke','Heußweg 37, 20255 Hamburg-Eimsbüttel','+49 40 409059','TBD','https://vita-apotheke-hh.de/','Contact via vita-apotheke-hh.de/kontakt-vita-apotheke/','Partial','Open 365 days, 8am-midnight.'),
('Hamburg','Apotheke','easyApotheke Altona','Altonaer Poststr. 15, Hamburg-Altona','+49 40 70298600','TBD','https://altona.easyapotheken.de/','TBD','Not started',None),
('Hamburg','Apotheke','Adler Apotheke Wandsbek','Wandsbeker Marktstrasse 73, 22041 Hamburg','+49 40 68942 0','TBD','https://adler-apotheke-hh.de/','TBD','Not started','Open daily until midnight.'),
('Hamburg','Apotheke','easyApotheke Wandsbek','Wandsbeker Marktstr. 1, Hamburg','+49 40 69463535','TBD','https://wandsbek.easyapotheken.de/','TBD','Not started',None),
('Hamburg','Apotheke','HamburgArznei-Apotheke Barmbek','Fuhlsbüttler Str. 107/109, 22305 Hamburg','+49 40 23935935','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Wald-Apotheke Barmbek','Rübenkamp 220, 22307 Hamburg','+49 40 63128570','TBD','https://www.wald-team.com/Seiten/Filialen/Barmbek.html','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke am Rathaus (Bergedorf)','Wentorfer Straße 35, 21029 Hamburg-Bergedorf','+49 40 7214147','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Bahnhof-Apotheke (Bergedorf)','Weidenbaumsweg 1, 21029 Hamburg-Bergedorf','+49 40 7213037','TBD','https://www.bahnhof-apotheke-bergedorf.de/','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke (Holtenklinker Straße)','Holtenklinker Str. 1A, 21029 Hamburg-Bergedorf','+49 40 7213015','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Hamburg','Apotheke','Löwen Apotheke Bergedorf','Hinterm Graben 31, 21029 Hamburg-Bergedorf','+49 40 7212200','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Mohnhof Apotheke','Hinterm Graben 26, 21029 Hamburg-Bergedorf','+49 40 72410000','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Pluspunkt Apotheke am Ärztehaus','Sachsentor 67, 21029 Hamburg-Bergedorf','+49 40 81974710','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Eppendorfer Marktplatz','Eppendorfer Marktplatz 2, 20251 Hamburg','+49 40 488778','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Eppendorf Botendienst','Martinistraße 64, 20251 Hamburg','+49 40 51326290','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Eppendorf','Eppendorfer Landstraße 84, 20249 Hamburg','+49 40 473420','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke Winterhude','Mexikoring 15, 22297 Hamburg','+49 40 630 2069','TBD','TBD','TBD','Not started',None),
('Hamburg','Apotheke','Marktplatz-Apotheke Winterhude','Winterhuder Marktplatz 21a, 22299 Hamburg','+49 40 473939','TBD','https://www.marktplatz-apotheke.com/','TBD','Not started',None),
('Hamburg','Apotheke','Apotheke (Alsterdorfer Straße)','Alsterdorfer Str. 6, 22299 Hamburg-Winterhude','+49 40 476005','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Hamburg','Apotheke','Apotheke (Jarrestraße)','Jarrestr. 42, 22303 Hamburg-Winterhude','+49 40 2708294','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Hamburg','Apotheke','City Apotheke Harburg','Lüneburger Straße 34, 21073 Hamburg-Harburg','+49 40 777030','TBD','https://city.weber-apotheken.de/','TBD','Not started',None),
('Hamburg','Apotheke','Arcaden Apotheke (Harburg)','Lüneburger Straße 45, 21073 Hamburg-Harburg','+49 40 30092121','TBD','https://arcaden.weber-apotheken.de/','TBD','Not started','Open daily 8am-midnight.'),
('Hamburg','Apotheke','Apotheke (Lüneburger Straße 18)','Lüneburger Straße 18, 21073 Hamburg-Harburg','+49 40 76755753','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
]

for r in rows:
    ws.append(r)

wb.save('ginger_shoc_outreach_tracker.xlsx')
count=0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]=='Apotheke': count+=1
print('Hamburg Apotheke total:', count)
