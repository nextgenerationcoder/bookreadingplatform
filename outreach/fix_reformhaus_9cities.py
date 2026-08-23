import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')

# (sheet, shop_name) -> updates dict
updates = {
    ('Stuttgart','Reformhaus (Klettpassage)'): dict(status='Excluded (chain)', website='https://www.vitalia-reformhaus.de/', contact='shop@vitalia-reformhaus.de (HQ)', note='CORRECTED: this location is actually VITALIA Reformhaus (Klettpassage 14+15), not independent — excluded as national chain.'),
    ('Düsseldorf','Reformhaus (Am Wehrhahn)'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/vita-nova/reformhaus-kaubisch-filialfinder', contact='info@reformhaus-kaubisch.de (same Kaubisch/Vita Nova chain contact as Essen entry)', note='CORRECTED: this location is Reformhaus Kaubisch (Vita Nova chain), not independent — excluded.'),
    ('Essen','Reformhaus (Rüttenscheider Straße)'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/vita-nova/reformhaus-kaubisch-ruettenscheid', contact='info@reformhaus-kaubisch.de (same Kaubisch/Vita Nova chain contact)', note='CORRECTED: this is also a Reformhaus Kaubisch (Vita Nova chain) branch, not independent — excluded.'),
    ('Hannover','Reformhaus (Ernst-August-Platz, Hauptbahnhof)'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/reformhaus-bacher/', contact='shop@reformhaus-bacher.de', note='CORRECTED: branded "betterlife", part of Reformhaus Bacher GmbH & Co. KG chain — excluded.'),
    ('Hannover','Reformhaus (Osterstraße)'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/reformhaus-bacher/', contact='shop@reformhaus-bacher.de', note='CORRECTED: also a Reformhaus Bacher chain branch — excluded.'),
    ('Hannover','Reformhaus (Schlägerstraße, Südstadt)'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/reformhaus-bacher/', contact='mail@reformhaus-bacher.de', note='CORRECTED: also a Reformhaus Bacher chain branch — excluded.'),
    ('Nuremberg','Reformhaus Mögeldorf'): dict(status='Excluded (chain)', website='https://reformhaus.de/blogs/vita-nova/reformhaus-seiler-nuernberg', contact='Contact via reformhaus.de/blogs/vita-nova/reformhaus-seiler-nuernberg — no direct email confirmed', note='CORRECTED: branded "Vita Nova Reformhaus Seiler" — chain, excluded.'),
}

for (sheet, name), u in updates.items():
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[2].value == name:
            row[8].value = u['status']
            row[6].value = u['website']
            row[7].value = u['contact']
            row[9].value = u['note']
            break

wb.save('ginger_shoc_outreach_tracker.xlsx')
print('corrections applied')
