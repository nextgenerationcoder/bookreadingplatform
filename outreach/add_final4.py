import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
headers = ['City', 'Category', 'Shop Name', 'Address', 'Phone', 'Rating', 'Website', 'Contact Email / Form', 'Outreach Status', 'Notes']

data = {
'Frankfurt am Main': [
    ('Frankfurt am Main','Reformhaus','Reformhaus Andersch ("Alles Andersch")','Glauburgstr. 77, 60318 Frankfurt-Nordend-West','+49 69 558861','TBD','https://www.alles-andersch.de/','achim@alles-andersch.de','Found','Independent, single location.'),
    ('Frankfurt am Main','Reformhaus','Reformhaus Freya (Schweizer Straße)','Schweizer Str. 18, 60594 Frankfurt-Sachsenhausen','+49 69 61990946','TBD','http://www.reformhaus-freya.de/','TBD — no public email found; phone contact only','Not started','Independent (Frankfurt-only KG, unrelated to Wiesbaden\'s separate Reformhaus FREYA KG despite matching name) — has a 2nd branch at Textorstr. 10, same city.'),
    ('Frankfurt am Main','Reformhaus','Liwell Reformhaus (Eschersheimer Landstr.)','Eschersheimer Landstr. 248, 60320 Frankfurt-Dornbusch','TBD','TBD','TBD','Contact via reformhaus.de — likely same "Liwell" multi-city brand already excluded in Wiesbaden','Excluded (chain)','Same Liwell brand family as Wiesbaden Herrmann/Liwell entries — treating as chain.'),
],
'Halle (Saale)': [
    ('Halle (Saale)','Reformhaus','BioRio Naturkost- und Reformhaus (= Reformhaus Quentin GmbH)','Große Ulrichstraße 16, 06108 Halle (Saale)','+49 345 2037611','TBD','http://www.biorio.de/','halle@biorio.de','Excluded (chain)','CORRECTED: this "BioRio" shop is actually operated by Reformhaus Quentin GmbH — the same company already excluded in Berlin (multi-city chain: Berlin + Halle confirmed). No independent alternative found yet in Halle — flagging as a gap for follow-up rather than fabricating an entry.'),
    ('Halle (Saale)','Reformhaus','[No independent Reformhaus identified yet]','','','','','','Not started','Only Reformhaus-branded shop surfaced in initial search was BioRio/Quentin (chain, excluded above). Needs a follow-up search (e.g. via neuform cooperative member directory) to find a true single-city independent.'),
],
'Magdeburg': [
    ('Magdeburg','Reformhaus','Reformhaus Inh. Rita Knackmuß','Große Diesdorfer Str. 229, 39108 Magdeburg-Stadtfeld','+49 391 7312224','TBD','TBD — no dedicated website found','TBD — no public email found; phone contact only','Not started','Independent, single owner-operator.'),
    ('Magdeburg','Reformhaus','Reformhaus Bacher (Allee-Center)','Ernst-Reuter-Allee 11, 39104 Magdeburg-Altstadt','TBD','TBD','https://reformhaus.de/blogs/reformhaus-bacher/','mail@reformhaus-bacher.de (same Bacher chain contact reused from Düsseldorf/Essen/Hannover)','Excluded (chain)','Same Reformhaus Bacher multi-city chain already excluded elsewhere.'),
],
'Freiburg im Breisgau': [
    ('Freiburg im Breisgau','Reformhaus','Drogerie-Reformhaus Friedrich','Uffhauser Str. 1, 79115 Freiburg-Haslach','+49 761 494341','TBD','TBD — no dedicated website found','TBD — no public email found; phone contact only','Not started','Independent, combined drugstore + Reformhaus.'),
    ('Freiburg im Breisgau','Reformhaus','Reformhaus (Habsburger Straße)','Habsburger Str. 105, 79104 Freiburg','TBD','TBD','TBD','TBD — chain affiliation unconfirmed, needs follow-up','Not started','Name/ownership not yet confirmed — verify by phone before outreach.'),
    ('Freiburg im Breisgau','Reformhaus','Reformhaus (Breisacher Straße)','Breisacher Str. 151, 79110 Freiburg','TBD','TBD','TBD','TBD — chain affiliation unconfirmed, needs follow-up','Not started','Name/ownership not yet confirmed — verify by phone before outreach.'),
],
}

for sheetname, rows in data.items():
    ws = wb.create_sheet(sheetname)
    ws.append(headers)
    for r in rows:
        ws.append(r)

wb.save('ginger_shoc_outreach_tracker.xlsx')
print('done', list(data.keys()))
print('total sheets:', len(wb.sheetnames))
