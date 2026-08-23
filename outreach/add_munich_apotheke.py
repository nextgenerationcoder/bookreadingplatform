import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
ws = wb['Munich']

rows = [
('Munich','Apotheke','Apotheke Rosenstraße','Rosenstr. 6, 80331 München-Altstadt','+49 89 23002700','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Adler-Apotheke','Dultstr. 1, 80331 München-Altstadt','+49 89 265477','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Frauenstraße','Frauenstr. 17, 80469 München-Altstadt','+49 89 225630','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Theatinerstraße','Theatinerstr. 45, 80333 München-Altstadt','+49 89 222163','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Neuhauser Straße','Neuhauser Str. 11, 80331 München-Altstadt','+49 89 5505070','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Eisenmannstraße','Eisenmannstr. 2, 80331 München-Altstadt','+49 89 32764934','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Sonnenstraße','Sonnenstr. 31, 80331 München-Altstadt','+49 89 593659','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Reichenbachstraße','Reichenbachstr. 9, 80469 München-Altstadt','+49 89 294429','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Falkenturmstraße','Falkenturmstr. 14, 80331 München-Altstadt','+49 89 221782','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke Maxvorstadt','Türkenstraße 42, 80799 München-Maxvorstadt','+49 89 284259','TBD','https://apothekemaxvorstadt.de/','info@apothekemaxvorstadt.de','Found','Same operator (Dr. Stefan Landshamer & Constanza Hagn) as Apotheke Schwabing Nord below.'),
('Munich','Apotheke','Apotheke Schwabing Nord','Ungererstraße 175, 80805 München-Schwabing','+49 89 52031954','TBD','https://apothekemaxvorstadt.de/','info@apothekemaxvorstadt.de','Found','Same operator as Apotheke Maxvorstadt above.'),
('Munich','Apotheke','Apotheke (Radlkoferstraße, Sendling)','Radlkoferstr. 2, 81373 München-Sendling','+49 89 45867885','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Valleystraße, Sendling)','Valleystr. 19, 81371 München-Sendling','+49 89 765211','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Thalkirchner Straße, Sendling)','Thalkirchner Str. 200, 81371 München-Sendling','+49 89 763041','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Dom-Pedro-Straße, Neuhausen)','Dom-Pedro-Str. 20, 80637 München-Neuhausen','+49 89 153316','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Gerner Straße, Neuhausen)','Gerner Str. 15, 80638 München-Neuhausen','+49 89 152174','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Volkartstraße, Neuhausen)','Volkartstr. 5, 80634 München-Neuhausen','+49 89 132036','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Lachner-Apotheke','Nymphenburger Str. 191, 80639 München','+49 89 160756','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Schloß-Apotheke','Romanplatz 5, 80639 München-Nymphenburg','+49 89 171884','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke (Grillparzerstraße, Haidhausen)','Grillparzerstr. 53, 81675 München-Haidhausen','+49 89 479367','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Einsteinstraße, Haidhausen)','Einsteinstr. 130, 81675 München-Haidhausen','+49 89 55060320','TBD','TBD','TBD','Not started','Formerly "Bienen-Apotheke Leuchtenbergring".'),
('Munich','Apotheke','Apotheke (Wörthstraße, Haidhausen)','Wörthstr. 3, 81667 München-Haidhausen','+49 89 4484107','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Bienen-Apotheke Giesing/Ramersdorf','Giesinger Bahnhofplatz 2, 81539 München','+49 89 21909 1330','TBD','TBD','TBD','Not started',None),
('Munich','Apotheke','Apotheke (Bajuwarenstraße, Trudering)','Bajuwarenstr. 50, 81825 München-Trudering','+49 89 422446','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Waldtruderinger Straße, Trudering)','Waldtruderinger Str. 67, 81827 München-Trudering','+49 89 4304815','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Truderinger Straße, Trudering)','Truderinger Str. 304A, 81825 München-Trudering','+49 89 421700','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Bäckerstraße, Pasing)','Bäckerstr. 4, 81241 München-Pasing','+49 89 844866','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (August-Exter-Straße, Pasing)','August-Exter-Str. 4, 81245 München-Pasing','+49 89 820930','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Munich','Apotheke','Apotheke (Landsberger Straße, Pasing)','Landsberger Str. 527, 81241 München-Pasing','+49 89 8200646','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
]

for r in rows:
    ws.append(r)

wb.save('ginger_shoc_outreach_tracker.xlsx')
count=0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]=='Apotheke': count+=1
print('Munich Apotheke total:', count)
