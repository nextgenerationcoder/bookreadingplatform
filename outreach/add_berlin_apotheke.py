import openpyxl
wb = openpyxl.load_workbook('ginger_shoc_outreach_tracker.xlsx')
ws = wb['Berlin (Pilot)']

rows = [
('Berlin','Apotheke','Apotheke am Kurfürstendamm','Kurfürstendamm 200, 10719 Berlin','+49 30 88778282','TBD','https://apo-kudamm.de/','Contact via apo-kudamm.de/Kontakt','Partial',None),
('Berlin','Apotheke','Brillant Apotheke','Kurfürstendamm 71, 10709 Berlin','+49 30 32766730','TBD','TBD','TBD','Not started',None),
('Berlin','Apotheke','Apotheke am Adenauerplatz','Kurfürstendamm 69, 10707 Berlin','+49 30 88917007','TBD','TBD','TBD','Not started',None),
('Berlin','Apotheke','Apotheke (Knobelsdorffstraße)','Knobelsdorffstraße 4, Berlin-Charlottenburg','+49 30 3214630','TBD','TBD','TBD','Not started','Business name not confirmed by search — address/phone from directory listing.'),
('Berlin','Apotheke','Apotheke (Bismarckstraße)','Bismarckstraße 89, Berlin-Charlottenburg','TBD','TBD','TBD','TBD','Not started','Business name not confirmed by search — address from directory listing.'),
('Berlin','Apotheke','Apotheke (Spandauer Damm)','Spandauer Damm 49, Berlin-Charlottenburg','+49 30 3218467','TBD','TBD','TBD','Not started','Business name not confirmed by search — address/phone from directory listing.'),
('Berlin','Apotheke','Apotheke (Mierendorffplatz)','Mierendorffplatz 12, Berlin-Charlottenburg','+49 30 3467080','TBD','TBD','TBD','Not started','Business name not confirmed by search — address/phone from directory listing.'),
('Berlin','Apotheke','Apotheke (Joachimsthaler Straße)','Joachimsthaler Straße 38, Berlin-Charlottenburg','+49 30 8826446','TBD','TBD','TBD','Not started','Business name not confirmed by search — address/phone from directory listing.'),
('Berlin','Apotheke','Apotheke zum Kreuz','Kastanienallee 2, 10435 Berlin','+49 30 4485181','TBD','TBD','TBD','Not started',None),
('Berlin','Apotheke','Kastanien-Apotheke','Kastanienallee 37-38, 13158 Berlin','+49 30 91207990','TBD','https://www.apotheken.de/13158/berlin/kastanien-apotheke/','TBD','Not started',None),
('Berlin','Apotheke','Sanimedius-Apotheke Kastanienallee (Dr. Ronald Clasen)','Kastanienallee, Berlin-Mitte','+49 30 4490244','TBD','TBD','TBD','Not started','Different branch from the Sanimedius Rosenthaler Platz entry above.'),
('Berlin','Apotheke','Kreuzberg-Apotheke','Mehringdamm 69, 10961 Berlin','+49 30 6937739','TBD','https://www.kreuzberg-apotheke-berlin-app.de/','TBD','Not started',None),
('Berlin','Apotheke','Reichenberger Apotheke','Reichenberger Str. 110, 10999 Berlin','+49 30 6126843','TBD','https://www.reichenberger-apotheke.de/','TBD','Not started',None),
('Berlin','Apotheke','Apotheke (Yorckstraße)','Yorckstr. 18, 10965 Berlin-Kreuzberg','+49 30 2158794','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Blücherstraße)','Blücherstr. 28, 10961 Berlin-Kreuzberg','+49 30 6919640','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Oranienstraße)','Oranienstr. 158, 10969 Berlin-Kreuzberg','+49 30 4202387','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Karl-Marx-Straße 214)','Karl-Marx-Str. 214, 12055 Berlin-Neukölln','TBD','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Hermannstraße)','Hermannstr. 144, 12051 Berlin-Neukölln','TBD','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Karl-Marx-Straße 239)','Karl-Marx-Str. 239, 12055 Berlin-Neukölln','TBD','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Sylter Apotheke','Berliner Str. 37, 10715 Berlin-Wilmersdorf','+49 30 8614211','TBD','https://www.sylter-apotheke.de/','Contact via sylter-apotheke.de/Kontakt','Partial',None),
('Berlin','Apotheke','Apotheke (Ollenhauer Str. 3-5)','Ollenhauer Str. 3-5, 13403 Berlin-Reinickendorf','+49 30 85622710','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Scharnweberstraße)','Scharnweberstr. 48, 13405 Berlin-Reinickendorf','+49 30 4121188','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Ollenhauerstraße 139)','Ollenhauerstr. 139, 13403 Berlin-Reinickendorf','0800 1008588','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Apotheke (Kopenhagener Straße)','Kopenhagener Str. 2, 13407 Berlin-Reinickendorf','+49 30 4955590','TBD','TBD','TBD','Not started','Business name not confirmed by search.'),
('Berlin','Apotheke','Altstadt-Apotheke','Breite Str. 20, 13597 Berlin-Spandau','+49 30 3332030','TBD','https://www.altstadt-apotheke-berlin.de/','TBD','Not started',None),
('Berlin','Apotheke','Apotheke am Bahnhof Spandau','Seegefelder Str. 16, 13597 Berlin-Spandau','+49 30 3334686','TBD','https://apothekebahnhofspandau.de/','TBD','Not started',None),
('Berlin','Apotheke','Herz Apotheke Spandau','13581 Berlin-Spandau','+49 30 33309393','TBD','https://www.herz-apotheke-spandau.de/','TBD','Not started',None),
]

for r in rows:
    ws.append(r)

wb.save('ginger_shoc_outreach_tracker.xlsx')
print('added', len(rows), 'rows')

# count total Apotheke for Berlin
count=0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]=='Apotheke': count+=1
print('Berlin Apotheke total:', count)
